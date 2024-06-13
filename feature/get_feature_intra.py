from feature.Intra_CrossTalk_class import IntraCrossTalk
from read_Excel import readExcel
from write_Excel import writeExcel
import openpyxl
from openpyxl.styles import Font

# Inter Cross-talk初始数据文件
Intra_Po_Cross_talk_file = '../Samples/Intra/Intra_positive.xlsx'
Intra_Ne_Cross_talk_file = '../Samples/Intra/Intra_negative.xlsx'
Positive_sheet, Po_rows, Po_cols = readExcel(Intra_Po_Cross_talk_file, 'Sheet1')
Negative_sheet, Ne_rows, Ne_cols = readExcel(Intra_Ne_Cross_talk_file, 'Sheet1')

# feature 文件路径
feature_path = r'D:\desktop\master4\Inter PTM cross-talk\results_feature'
alphafold_enm_feature_path = r'D:\desktop\master4\Inter PTM cross-talk\ENM_Alpha\results'
# enm算出来的结构特征 Alphafold
enm_anm_cancer_cc_alphafold_path = alphafold_enm_feature_path + r'\anm\cc'  # 5_per,5_per_20_per, 20_per_50_per, greater_60_per, top3
enm_anm_cancer_prs_alphafold_path = alphafold_enm_feature_path + r'\anm\prs'  # effectiveness, prs, sensitivity
enm_anm_cancer_sq_alphafold_path = alphafold_enm_feature_path + r'\anm\sq'  # sq
enm_anm_cancer_stiffness_alphafold_path = alphafold_enm_feature_path + r'\anm\stiffness'  # stiffness
enm_gnm_cancer_cc_alphafold_path = alphafold_enm_feature_path + r'\gnm\cc'  # 5_per,5_per_20_per, 20_per_50_per, greater_60_per, top3
enm_gnm_cancer_eigenvector_alphafold_path = alphafold_enm_feature_path + r'\gnm\eigenvector'  # eigenvector_20, eigenvector_all, eigenvector_top3
enm_gnm_cancer_prs_alphafold_path = alphafold_enm_feature_path + r'\gnm\prs'  # effectiveness, prs, sensitivity
enm_gnm_cancer_sq_alphafold_path = alphafold_enm_feature_path + r'\gnm\sq'  # sq

# uniprot算出来的序列特征evol
evol_dirinfo = feature_path + r'\evol\dirinfo'
evol_entropy = feature_path + r'\evol\entropy'
evol_mifc = feature_path + r'\evol\mifc'
evol_mifn = feature_path + r'\evol\mifn'
evol_mutinfo = feature_path + r'\evol\mutinfo'
evol_occupancy = feature_path + r'\evol\occupancy'
evol_omes = feature_path + r'\evol\omes'
evol_sca = feature_path + r'\evol\sca'

# Nacen计算出来的特征
alphafold_nacen_path = r'D:/desktop/master4/Inter PTM cross-talk/NACEN/NACEN_alphafold_feature/'


def get_feature(sheet, rows, cols, featurefile):
    print(sheet, rows, cols)  # <xlrd.sheet.Sheet object at 0x0000024792E09EC8> 10722 21

    Intra_titles = ['Pro_name', 'UniprotId', 'UniprotSite1', "Site1Type", 'UniprotSite2', 'Site2Type',
                    'unw_degree_alphafold_nacen', 'unw_betweenness_alphafold_nacen', 'unw_closeness_alphafold_nacen',
                    'w_degree_alphafold_nacen', 'w_betweeness_alphafold_nacen', 'w_closeness_alphafold_nacen',
                    'anm_cc_20_per_50_per_alphafold', 'anm_cc_5_per_alphafold', 'anm_cc_5_per_20_per_alphafold',
                    'anm_cc_greater_60_per_alphafold',
                    'anm_cc_top3_alphafold', 'anm_effectiveness_all_alphafold', 'anm_prs_all_alphafold',
                    'anm_sensitivity_all_alphafold', 'anm_sq_all_alphafold',
                    'anm_stiffness_alphafold', 'gnm_cc_20_per_50_per_alphafold', 'gnm_cc_5_per_alphafold',
                    'gnm_cc_5_per_20_per_alphafold',
                    'gnm_cc_greater_60_per_alphafold', 'gnm_cc_top3_alphafold', 'gnm_eigenvectors_20_alphafold',
                    'gnm_eigenvectors_all_alphafold',
                    'gnm_eigenvectors_top3_alphafold', 'gnm_effectiveness_all_alphafold', 'gnm_prs_all_alphafold',
                    'gnm_sensitivity_all_alphafold',
                    'gnm_sq_all_alphafold',
                    'evol_dirinfo', 'evol_entropy', 'evol_mifc', 'evol_mifn', 'evol_mutinfo', 'evol_occupancy',
                    'evol_omes', 'evol_sca']
    Intra_features_name = ['unw_degree_alphafold_nacen', 'unw_betweenness_alphafold_nacen',
                           'unw_closeness_alphafold_nacen',
                           'w_degree_alphafold_nacen', 'w_betweeness_alphafold_nacen', 'w_closeness_alphafold_nacen',
                           'anm_cc_20_per_50_per_alphafold', 'anm_cc_5_per_alphafold', 'anm_cc_5_per_20_per_alphafold',
                           'anm_cc_greater_60_per_alphafold',
                           'anm_cc_top3_alphafold', 'anm_effectiveness_all_alphafold', 'anm_prs_all_alphafold',
                           'anm_sensitivity_all_alphafold', 'anm_sq_all_alphafold',
                           'anm_stiffness_alphafold', 'gnm_cc_20_per_50_per_alphafold', 'gnm_cc_5_per_alphafold',
                           'gnm_cc_5_per_20_per_alphafold',
                           'gnm_cc_greater_60_per_alphafold', 'gnm_cc_top3_alphafold', 'gnm_eigenvectors_20_alphafold',
                           'gnm_eigenvectors_all_alphafold',
                           'gnm_eigenvectors_top3_alphafold', 'gnm_effectiveness_all_alphafold',
                           'gnm_prs_all_alphafold',
                           'gnm_sensitivity_all_alphafold',
                           'gnm_sq_all_alphafold',
                           'evol_dirinfo', 'evol_entropy', 'evol_mifc', 'evol_mifn', 'evol_mutinfo', 'evol_occupancy',
                           'evol_omes', 'evol_sca']

    for row in range(1, rows + 1):
        # for row in range(1, 5):
        if row == 1:
            Intra_Cross_talk_excel = openpyxl.Workbook()
            sheet1 = Intra_Cross_talk_excel.create_sheet('Sheet', 0)
            for i in range(1, len(Intra_titles) + 1):
                sheet1.cell(row, i, value=Intra_titles[i - 1])
        else:
            features_all = []
            features_Pro1 = []
            features_Pro2 = []
            print(row)
            print(sheet.row_values(row - 1))
            Pro_name = str(sheet.row_values(row - 1)[0])
            UniprotId = str(sheet.row_values(row - 1)[1])
            UniprotSite1 = str(sheet.row_values(row - 1)[2])
            Site1Type = str(sheet.row_values(row - 1)[3])
            UniprotSite2 = str(sheet.row_values(row - 1)[4])
            Site2Type = str(sheet.row_values(row - 1)[5])
            Intra_Cross_talk = IntraCrossTalk(Pro_name, UniprotId, UniprotSite1, Site1Type,
                                              UniprotSite2, Site2Type)
            print('------nacen-----')
            Intra_Cross_talk.pdbchain_feature_nacen(alphafold_nacen_path)
            print('-------enm-------')
            # alphafold
            Intra_Cross_talk.pdbchain_feature_anm_cc(enm_anm_cancer_cc_alphafold_path)
            Intra_Cross_talk.pdbchain_feature_anm_prs(enm_anm_cancer_prs_alphafold_path)
            Intra_Cross_talk.pdbchain_feature_anm_sq(enm_anm_cancer_sq_alphafold_path)
            Intra_Cross_talk.pdbchain_feature_anm_stiffness(enm_anm_cancer_stiffness_alphafold_path)
            Intra_Cross_talk.pdbchain_feature_gnm_cc(enm_gnm_cancer_cc_alphafold_path)
            Intra_Cross_talk.pdbchain_feature_gnm_eigenvector(enm_gnm_cancer_eigenvector_alphafold_path)
            Intra_Cross_talk.pdbchain_feature_gnm_prs(enm_gnm_cancer_prs_alphafold_path)
            Intra_Cross_talk.pdbchain_feature_gnm_sq(enm_gnm_cancer_sq_alphafold_path)

            # print(Ptm_mutation.UniprotPosition)
            Intra_Cross_talk.uninprot_feature_evol_dirinfo(evol_dirinfo)
            Intra_Cross_talk.uninprot_feature_evol_entropy(evol_entropy)
            Intra_Cross_talk.uninprot_feature_evol_mifc(evol_mifc)
            Intra_Cross_talk.uninprot_feature_evol_mifn(evol_mifn)
            Intra_Cross_talk.uninprot_feature_evol_mutinfo(evol_mutinfo)
            Intra_Cross_talk.uninprot_feature_evol_occupancy(evol_occupancy)
            Intra_Cross_talk.uninprot_feature_evol_omes(evol_omes)
            Intra_Cross_talk.uninprot_feature_evol_sca(evol_sca)
            features_Pro1.extend([Pro_name, UniprotId, UniprotSite1, Site1Type])
            features_Pro2.extend([Pro_name, UniprotId, UniprotSite2, Site2Type])
            features_all.extend(
                [Pro_name, UniprotId, UniprotSite1, Site1Type,
                 UniprotSite2, Site2Type])
            print('NACEN_features: ', len(Intra_Cross_talk.pdbchain1_nacen_feature),
                  len(Intra_Cross_talk.pdbchain2_nacen_feature))
            print('Enm_features: ', len(Intra_Cross_talk.pdbchain1_enm_feature),
                  len(Intra_Cross_talk.pdbchain2_enm_feature))
            print('Uniprot_features: ', len(Intra_Cross_talk.uniprot1_evol_feature),
                  len(Intra_Cross_talk.uniprot2_evol_feature))
            print('All_features: ', len(Intra_Cross_talk.features))

            print(Intra_Cross_talk.features)
            for feature_name in Intra_features_name:
                features_all.append(Intra_Cross_talk.features[feature_name])
            print(features_all)
            sheet1.append(features_all)

    Intra_Cross_talk_excel.save(featurefile)


get_feature(Positive_sheet, Po_rows, Po_cols, './Intra/Positive_features0701.xlsx')
get_feature(Negative_sheet, Ne_rows, Ne_cols, './Intra/Negative_features0701.xlsx')
